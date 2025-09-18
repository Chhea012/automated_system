from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from app import db
from app.models.contract import Contract
from app.models.department import Department
from app.models.user import User
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/contracts')
@login_required
def contract_report():
    # Filters
    department_id = request.args.get('department_id', 'all')
    search = request.args.get('search', '').strip().lower()
    month_year = request.args.get('month_year', datetime.now().strftime('%B %Y'))
    sort = request.args.get('sort', 'contract_number_asc')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    view_mode = request.args.get('view_mode', 'monthly')
    day_filter = request.args.get('day_filter', 'All')

    # Parse month_year
    try:
        month_name, year_str = month_year.split()
        month_num = datetime.strptime(month_name, '%B').month
        year = int(year_str)
        # Convert full month name to abbreviated for display
        month_abbr = datetime.strptime(month_name, '%B').strftime('%b')
        month_year_display = f"{month_abbr} {year}"
    except (ValueError, AttributeError):
        month_num = datetime.now().month
        year = datetime.now().year
        month_year = datetime.now().strftime('%B %Y')
        month_year_display = datetime.now().strftime('%b %Y')

    # Day map for dayofweek
    day_map = {'Mon': 2, 'Tue': 3, 'Wed': 4, 'Thu': 5, 'Fri': 6, 'Sat': 7, 'Sun': 1}

    # Base query for table (filtered by month and year)
    query = Contract.query.filter(Contract.deleted_at == None)\
                          .outerjoin(Contract.user)\
                          .filter(db.extract('year', Contract.created_at) == year)\
                          .filter(db.extract('month', Contract.created_at) == month_num)

    if department_id != 'all':
        query = query.filter(Contract.user.has(department_id=department_id))

    if search:
        query = query.filter(
            (Contract.project_title.ilike(f'%{search}%')) |
            (Contract.party_b_signature_name.ilike(f'%{search}%')) |
            (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
        )

    if day_filter != 'All':
        dow = day_map.get(day_filter)
        if dow:
            query = query.filter(db.func.dayofweek(Contract.created_at) == dow)

    # Sorting
    if sort == 'contract_number_desc':
        query = query.order_by(Contract.contract_number.desc())
    elif sort == 'project_title_asc':
        query = query.order_by(Contract.project_title.asc())
    elif sort == 'project_title_desc':
        query = query.order_by(Contract.project_title.desc())
    else:
        query = query.order_by(Contract.contract_number.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    contracts = [contract.to_dict() for contract in pagination.items]

    # Totals
    total_contracts = query.count()
    departments = Department.query.all()
    department_totals = {}
    if department_id == 'all':
        for dept in departments:
            dept_query = Contract.query.filter(Contract.deleted_at == None)\
                                       .outerjoin(Contract.user)\
                                       .filter(Contract.user.has(department_id=dept.id))\
                                       .filter(db.extract('year', Contract.created_at) == year)\
                                       .filter(db.extract('month', Contract.created_at) == month_num)
            if search:
                dept_query = dept_query.filter(
                    (Contract.project_title.ilike(f'%{search}%')) |
                    (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                    (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                )
            if day_filter != 'All':
                dow = day_map.get(day_filter)
                if dow:
                    dept_query = dept_query.filter(db.func.dayofweek(Contract.created_at) == dow)
            department_totals[dept.name] = dept_query.count()
    else:
        selected_dept = Department.query.get(department_id)
        if selected_dept:
            dept_query = Contract.query.filter(Contract.deleted_at == None)\
                                       .outerjoin(Contract.user)\
                                       .filter(Contract.user.has(department_id=department_id))\
                                       .filter(db.extract('year', Contract.created_at) == year)\
                                       .filter(db.extract('month', Contract.created_at) == month_num)
            if search:
                dept_query = dept_query.filter(
                    (Contract.project_title.ilike(f'%{search}%')) |
                    (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                    (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                )
            if day_filter != 'All':
                dow = day_map.get(day_filter)
                if dow:
                    dept_query = dept_query.filter(db.func.dayofweek(Contract.created_at) == dow)
            department_totals[selected_dept.name] = dept_query.count()

    # Unique month_years (abbreviated for display)
    unique_months = db.session.query(
        db.func.distinct(db.func.date_format(Contract.created_at, '%M %Y'))
    ).filter(Contract.created_at != None).all()
    unique_months = [datetime.strptime(m[0], '%B %Y').strftime('%b %Y') for m in unique_months if m[0]]

    # Chart Data
    if view_mode == 'monthly':
        # Monthly counts for the year
        monthly_query = db.session.query(
            db.func.month(Contract.created_at),
            db.func.count(Contract.id)
        ).filter(
            Contract.deleted_at == None,
            db.extract('year', Contract.created_at) == year
        )
        if department_id != 'all':
            monthly_query = monthly_query.filter(Contract.user.has(department_id=department_id))
        if search:
            monthly_query = monthly_query.outerjoin(Contract.user).filter(
                (Contract.project_title.ilike(f'%{search}%')) |
                (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
            )
        if day_filter != 'All':
            dow = day_map.get(day_filter)
            if dow:
                monthly_query = monthly_query.filter(db.func.dayofweek(Contract.created_at) == dow)
        monthly_counts = monthly_query.group_by(
            db.func.month(Contract.created_at)
        ).all()

        monthly_dict = {m: c for m, c in monthly_counts}
        chart_labels = [
            'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
        ]
        chart_values = [monthly_dict.get(i, 0) for i in range(1, 13)]
        chart_title = f"Contracts Overview (Jan - Dec {year})"
    else:
        # Weekly view: Aggregated by day of week (Monday to Sunday) for the selected month
        first_day = datetime(year, month_num, 1)
        last_day = (first_day + timedelta(days=31)).replace(day=1) - timedelta(days=1)

        day_query = db.session.query(
            db.func.dayofweek(Contract.created_at),
            db.func.count(Contract.id)
        ).filter(
            Contract.deleted_at == None,
            Contract.created_at >= first_day,
            Contract.created_at < first_day + timedelta(days=32 - first_day.day)  # Strict < next month first day
        )
        if department_id != 'all':
            day_query = day_query.filter(Contract.user.has(department_id=department_id))
        if search:
            day_query = day_query.outerjoin(Contract.user).filter(
                (Contract.project_title.ilike(f'%{search}%')) |
                (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
            )
        if day_filter != 'All':
            dow = day_map.get(day_filter)
            if dow:
                day_query = day_query.filter(db.func.dayofweek(Contract.created_at) == dow)
        day_counts = day_query.group_by(
            db.func.dayofweek(Contract.created_at)
        ).all()

        # DAYOFWEEK: 1=Sunday, 2=Monday, ..., 7=Saturday
        dow_dict = {int(dow): count for dow, count in day_counts}

        # Map to Monday (2) to Sunday (1)
        chart_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        chart_values = [
            dow_dict.get(2, 0),  # Monday
            dow_dict.get(3, 0),  # Tuesday
            dow_dict.get(4, 0),  # Wednesday
            dow_dict.get(5, 0),  # Thursday
            dow_dict.get(6, 0),  # Friday
            dow_dict.get(7, 0),  # Saturday
            dow_dict.get(1, 0)   # Sunday
        ]

        # Ensure at least one data point
        if sum(chart_values) == 0:
            chart_labels = ['No Data']
            chart_values = [0]
        chart_title = f"Contracts by Day of Week ({month_year_display})"

    # Pie chart (department distribution for the month)
    pie_labels = list(department_totals.keys())
    pie_values = list(department_totals.values())
    pie_title = "Contracts by Department"
    if day_filter != 'All':
        pie_title += f" (on {day_filter}s)"
    if department_id != 'all':
        selected_dept = Department.query.get(department_id)
        if selected_dept:
            pie_title += f" ({selected_dept.name})"

    return render_template('reports/index.html',
                           contracts=contracts,
                           pagination=pagination,
                           departments=departments,
                           department_id=department_id,
                           search=search,
                           month_year=month_year,  # Keep full month name for URL/filter consistency
                           month_year_display=month_year_display,  # Abbreviated for display
                           sort=sort,
                           per_page=per_page,
                           total_contracts=total_contracts,
                           department_totals=department_totals,
                           unique_months=unique_months,
                           chart_labels=chart_labels,
                           chart_values=chart_values,
                           pie_labels=pie_labels,
                           pie_values=pie_values,
                           chart_title=chart_title,
                           pie_title=pie_title,
                           view_mode=view_mode,
                           day_filter=day_filter)

@reports_bp.route('/export_contracts_excel')
@login_required
def export_contracts_excel():
    department_id = request.args.get('department_id', 'all')
    month_year = request.args.get('month_year', datetime.now().strftime('%B %Y'))
    search = request.args.get('search', '').strip().lower()
    day_filter = request.args.get('day_filter', 'All')

    try:
        month_name, year_str = month_year.split()
        month_num = datetime.strptime(month_name, '%B').month
        year = int(year_str)
    except (ValueError, AttributeError):
        month_num = datetime.now().month
        year = datetime.now().year

    # Day map for dayofweek
    day_map = {'Mon': 2, 'Tue': 3, 'Wed': 4, 'Thu': 5, 'Fri': 6, 'Sat': 7, 'Sun': 1}

    query = Contract.query.filter(Contract.deleted_at == None)\
                          .outerjoin(Contract.user)\
                          .filter(db.extract('year', Contract.created_at) == year)\
                          .filter(db.extract('month', Contract.created_at) == month_num)

    if department_id != 'all' and department_id != 'current':
        query = query.filter(Contract.user.has(department_id=department_id))

    if search:
        query = query.filter(
            (Contract.project_title.ilike(f'%{search}%')) |
            (Contract.party_b_signature_name.ilike(f'%{search}%')) |
            (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
        )

    if day_filter != 'All':
        dow = day_map.get(day_filter)
        if dow:
            query = query.filter(db.func.dayofweek(Contract.created_at) == dow)

    contracts = query.all()
    data = [{
        'Number of Contract': c.contract_number,
        'Project Title': c.project_title,
        'Department': c.user.department.name if c.user and c.user.department else 'N/A',
        'Manager': c.user.username if c.user else 'N/A',
        'Contract Date': c.formatted_created_at,
        'Party B': c.party_b_signature_name
    } for c in contracts]

    df = pd.DataFrame(data)

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Report"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
        cell.fill = fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value='Total Contracts').font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(contracts)).font = Font(bold=True)

    wb.save(output)
    output.seek(0)

    filename = f"Contract_Report_{month_year.replace(' ', '_')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)